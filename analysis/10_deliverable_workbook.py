#!/usr/bin/env python3
"""
Build COCHRANE_CLIENT_LIST_EXPLAINED.xlsx from the release CSVs.

One workbook a non-technical reader can open and follow:
  START HERE          the two populations, the four cohorts, the number to remember, a 30-second script
  HOW PEOPLE ARE COUNTED  the logic, step by step
  HEADLINE            every figure is a live formula over the data sheets, next to the accepted value
  CLIENT LIST         one row per person, plain-English headers, real dates
  WAITLIST ACTIVITY   one row per waitlist spell
  PLACEMENT ACTIVITY  one row per Type A/B admission
  DATA DICTIONARY     plain label -> technical field -> definition
  QA CHECKS           live formulas, PASS/FAIL, one overall cell
No number is typed: values come from the CSVs, summaries are formulas.
"""
import csv, os, argparse, datetime as dt
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

FYES = [2022, 2023, 2024, 2025, 2026]
INK, ACC, MUT = "15201C", "058C72", "5C6862"
HFILL = PatternFill("solid", fgColor="F1F5F3"); KFILL = PatternFill("solid", fgColor="E6F3EF"); WARN = PatternFill("solid", fgColor="FFF4E5")
thin = Side(style="thin", color="D9DEDA"); BOX = Border(top=thin, bottom=thin, left=thin, right=thin)
B = Font(bold=True, color=INK); T = Font(bold=True, size=16, color=INK); H = Font(bold=True, size=12, color=ACC); N = Font(color=INK); M = Font(color=MUT, italic=True)
WRAP = Alignment(wrap_text=True, vertical="top")

def read(p): return list(csv.DictReader(open(p)))
def d(s): return dt.datetime.strptime(s[:10], "%Y-%m-%d").date() if s else None
def yn(v): return "Yes" if v == "1" else "No"
def num(v): return int(v) if v not in ("", None) else None

STATUS = {"incident demand FY2022-FY2026": "New demand in FY2022-FY2026",
          "pre-window demand (carry-in)": "Approved before April 2021; activity in the period (carry-in)",
          "prior residential care before FY2022 (activity only)": "In residential care before April 2021; activity in the period",
          "activity only (demand before FY2022 or not new demand)": "Earlier or non-new demand; activity in the period"}
COH = {"A": "A - Town of Cochrane resident, first placed in a Cochrane facility", "B": "B - Not a Town resident, placed in a Cochrane facility",
       "C": "C - Town of Cochrane resident, first placed outside Cochrane", "D": "D - Town of Cochrane resident, no placement observed by 31 March 2026"}
DCLS = {"D1": "D1 - still on the waitlist on 31 March 2026", "D2": "D2 - died before any placement", "D3": "D3 - left the list, no placement observed"}
def group(p):
    if p["COHORT"]: return COH[p["COHORT"]]
    pop = p["POPULATION"]
    if pop.startswith("X1"): return "New demand, placed in Cochrane, residence unresolved"
    if pop.startswith("X2"): return "New demand, not a Town resident, rated for a Cochrane site, first placed elsewhere or unplaced"
    if pop.startswith("X3"): return "New demand, residence unresolved, rated for a Cochrane site, first placed elsewhere or unplaced"
    if pop.startswith("X4"): return "New demand, later move into a Cochrane facility after a first placement elsewhere"
    return "Activity only - " + STATUS.get(p["ACTIVITY_STATUS"].split(":")[0], p["ACTIVITY_STATUS"])
RES = {"Town": "Town of Cochrane", "Cochrane catchment": "Cochrane catchment (Rocky View area)", "non-Town": "Not a Cochrane-area resident", "unresolved": "Unresolved"}

# ── column specs: (plain header, technical field, value function) ────────────
def person_cols():
    return [
        ("Study ID", "STUDY_ID", lambda p: p["STUDY_ID"]),
        ("New demand FY2022-FY2026?", "INCIDENT_DEMAND_SCOPE", lambda p: yn(p["INCIDENT_DEMAND_SCOPE"])),
        ("Cochrane activity FY2022-FY2026?", "CONSULTANT_ACTIVITY_SCOPE", lambda p: yn(p["CONSULTANT_ACTIVITY_SCOPE"])),
        ("Why the person is in the list", "ACTIVITY_STATUS", lambda p: STATUS.get(p["ACTIVITY_STATUS"].split(":")[0], p["ACTIVITY_STATUS"])),
        ("Cohort", "COHORT", lambda p: p["COHORT"]),
        ("Group (in words)", "POPULATION / COHORT", group),
        ("D outcome", "D_CLASS", lambda p: DCLS.get(p["D_CLASS"][:2], "") if p["COHORT"] == "D" else ""),
        ("Reference date", "DEMAND_DT or ACTIVITY_ANCHOR_DT", lambda p: d(p["DEMAND_DT"]) if p["INCIDENT_DEMAND_SCOPE"] == "1" else d(p["ACTIVITY_ANCHOR_DT"])),
        ("What the reference date is", "ATTRIBUTE_ANCHOR", lambda p: "Demand event (first Type A/B approval or admission)" if p["INCIDENT_DEMAND_SCOPE"] == "1" else "First activity in the period: " + p["ACTIVITY_ANCHOR_TYPE"]),
        ("Demand fiscal year", "DEMAND_FYE", lambda p: num(p["DEMAND_FYE"])),
        ("Demand event type", "DEMAND_EVENT_TYPE", lambda p: p["DEMAND_EVENT_TYPE"]),
        ("First waitlist date", "FIRST_WAITLIST_APPEARANCE", lambda p: d(p["FIRST_WAITLIST_APPEARANCE"])),
        ("Last seen on waitlist", "LAST_SEEN_ON_LIST", lambda p: d(p["LAST_SEEN_ON_LIST"])),
        ("Still on list at 31 Mar 2026?", "ON_LIST_AT_FOLLOWUP", lambda p: yn(p["ON_LIST_AT_FOLLOWUP"])),
        ("Age at reference date", "AGE_AT_ANCHOR", lambda p: num(p["AGE_AT_ANCHOR"])),
        ("Age group", "AGE_GROUP_AT_ANCHOR", lambda p: p["AGE_GROUP_AT_ANCHOR"]),
        ("Sex", "SEX", lambda p: p["SEX"]),
        ("Sex source", "SEX_SOURCE", lambda p: p["SEX_SOURCE"]),
        ("Residence", "RESIDENCE_CLASS", lambda p: RES.get(p["RESIDENCE_CLASS"], p["RESIDENCE_CLASS"])),
        ("Residence community", "RESIDENCE_COMMUNITY", lambda p: p["RESIDENCE_COMMUNITY"]),
        ("Residence address basis", "RESIDENCE_REFERENCE_TYPE / _FYE", lambda p: (p["RESIDENCE_REFERENCE_TYPE"] + (" FY" + p["RESIDENCE_REFERENCE_FYE"] if p["RESIDENCE_REFERENCE_FYE"] else "")) if p["RESIDENCE_REFERENCE_TYPE"] else ""),
        ("Origin setting", "ORIGIN_SETTING", lambda p: p["ORIGIN_SETTING"]),
        ("Origin detail", "ORIGIN_SETTING_DETAIL", lambda p: p["ORIGIN_SETTING_DETAIL"]),
        ("Origin tied on entry day?", "ORIGIN_CONFLICT_FLAG", lambda p: yn(p["ORIGIN_CONFLICT_FLAG"])),
        ("Rated for a Cochrane site?", "RATED_FOR_COCHRANE_SITE_FLAG", lambda p: yn(p["RATED_FOR_COCHRANE_SITE_FLAG"])),
        ("Cochrane sites rated", "COCHRANE_SITES_RATED", lambda p: p["COCHRANE_SITES_RATED"]),
        ("Number of sites rated", "N_SITES_RATED", lambda p: num(p["N_SITES_RATED"])),
        ("Most often observed rated site", "MOST_FREQUENTLY_OBSERVED_RATED_SITE", lambda p: p["MOST_FREQUENTLY_OBSERVED_RATED_SITE"]),
        ("First placement date", "FIRST_PLACEMENT_DT", lambda p: d(p["FIRST_PLACEMENT_DT"])),
        ("Placement fiscal year", "PLACEMENT_FYE", lambda p: num(p["PLACEMENT_FYE"])),
        ("First placement site", "FIRST_PLACEMENT_SITE", lambda p: p["FIRST_PLACEMENT_SITE"]),
        ("First placement type", "FIRST_PLACEMENT_STREAM", lambda p: p["FIRST_PLACEMENT_STREAM"]),
        ("First placement in Cochrane?", "FIRST_PLACEMENT_IN_COCHRANE", lambda p: yn(p["FIRST_PLACEMENT_IN_COCHRANE"]) if p["FIRST_PLACEMENT_DT"] else ""),
        ("Any Cochrane admission in period?", "ACTIVITY_COCHRANE_ADMISSION", lambda p: yn(p["ACTIVITY_COCHRANE_ADMISSION"])),
        ("Days from demand to first placement", "DAYS_TO_PLACEMENT", lambda p: num(p["DAYS_TO_PLACEMENT"])),
        ("Days waiting as of 31 Mar 2026 (still waiting only)", "DAYS_WAITING_AS_OF_FOLLOWUP", lambda p: num(p["DAYS_WAITING_AS_OF_FOLLOWUP"])),
    ]
def wait_cols():
    return [("Study ID", "STUDY_ID", lambda w: w["STUDY_ID"]), ("Spell number for this person", "SPELL_SEQ_FOR_PERSON", lambda w: num(w["SPELL_SEQ_FOR_PERSON"])),
            ("List entry date", "LIST_ENTRY_DT", lambda w: d(w["LIST_ENTRY_DT"])), ("List entry fiscal year", "LIST_ENTRY_FYE", lambda w: num(w["LIST_ENTRY_FYE"])),
            ("Last seen on list", "LIST_LAST_SEEN_DT", lambda w: d(w["LIST_LAST_SEEN_DT"])), ("Days observed in spell", "DAYS_OBSERVED", lambda w: num(w["DAYS_OBSERVED"])),
            ("Care type at entry", "CARE_STREAM_AT_ENTRY", lambda w: w["CARE_STREAM_AT_ENTRY"]), ("Origin setting", "ORIGIN_SETTING", lambda w: w["ORIGIN_SETTING"]), ("Origin detail", "ORIGIN_SETTING_DETAIL", lambda w: w["ORIGIN_SETTING_DETAIL"]),
            ("Origin tied on entry day?", "ORIGIN_CONFLICT_FLAG", lambda w: yn(w["ORIGIN_CONFLICT_FLAG"])), ("First approval date in spell", "FIRST_APPROVED_DT_IN_SPELL", lambda w: d(w["FIRST_APPROVED_DT_IN_SPELL"])),
            ("Rated a Cochrane site in this spell?", "RATED_COCHRANE_IN_SPELL", lambda w: yn(w["RATED_COCHRANE_IN_SPELL"])), ("Already on list at 1 Apr 2021?", "LEFT_TRUNCATED", lambda w: yn(w["LEFT_TRUNCATED"])),
            ("Still on list at 31 Mar 2026?", "ON_LIST_AT_FOLLOWUP", lambda w: yn(w["ON_LIST_AT_FOLLOWUP"])), ("Person's residence", "PERSON_RESIDENCE_CLASS", lambda w: RES.get(w["PERSON_RESIDENCE_CLASS"], w["PERSON_RESIDENCE_CLASS"])),
            ("Why the person is in the list", "PERSON_ACTIVITY_STATUS", lambda w: STATUS.get(w["PERSON_ACTIVITY_STATUS"], w["PERSON_ACTIVITY_STATUS"])), ("Person's cohort", "PERSON_COHORT", lambda w: w["PERSON_COHORT"])]
def event_cols():
    return [("Study ID", "STUDY_ID", lambda e: e["STUDY_ID"]), ("Admission date", "ADMISSION_DT", lambda e: d(e["ADMISSION_DT"])), ("Placement fiscal year", "PLACEMENT_FYE", lambda e: num(e["PLACEMENT_FYE"])),
            ("Facility", "PLACEMENT_SITE", lambda e: e["PLACEMENT_SITE"]), ("Care type", "CARE_STREAM", lambda e: e["CARE_STREAM"]), ("Cochrane facility?", "PLACEMENT_IN_COCHRANE", lambda e: yn(e["PLACEMENT_IN_COCHRANE"])),
            ("Admitted from (setting)", "ORIGIN_SETTING", lambda e: e["ORIGIN_SETTING"]), ("Admitted from (detail)", "ORIGIN_SETTING_DETAIL", lambda e: e["ORIGIN_SETTING_DETAIL"]),
            ("First placement of a new-demand person?", "IS_FIRST_PLACEMENT_INCIDENT", lambda e: yn(e["IS_FIRST_PLACEMENT_INCIDENT"])), ("Person's first admission in the period?", "IS_FIRST_PLACEMENT_IN_WINDOW", lambda e: yn(e["IS_FIRST_PLACEMENT_IN_WINDOW"])),
            ("Admission number for this person", "EVENT_SEQ_FOR_PERSON", lambda e: num(e["EVENT_SEQ_FOR_PERSON"])), ("Person's residence", "PERSON_RESIDENCE_CLASS", lambda e: RES.get(e["PERSON_RESIDENCE_CLASS"], e["PERSON_RESIDENCE_CLASS"])),
            ("Person's residence community", "PERSON_RESIDENCE_COMMUNITY", lambda e: e["PERSON_RESIDENCE_COMMUNITY"]), ("Why the person is in the list", "PERSON_ACTIVITY_STATUS", lambda e: STATUS.get(e["PERSON_ACTIVITY_STATUS"], e["PERSON_ACTIVITY_STATUS"])),
            ("Person's cohort", "PERSON_COHORT", lambda e: e["PERSON_COHORT"])]

def data_sheet(wb, name, rows, cols, widths):
    ws = wb.create_sheet(name)
    for j, (h, _, _) in enumerate(cols, 1):
        c = ws.cell(1, j, h); c.font = B; c.fill = HFILL; c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = BOX
    for i, r in enumerate(rows, 2):
        for j, (_, _, f) in enumerate(cols, 1):
            v = f(r); c = ws.cell(i, j, v)
            if isinstance(v, dt.date): c.number_format = "yyyy-mm-dd"
    for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 42; ws.freeze_panes = "B2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    return ws
def L(cols, header): return get_column_letter([h for h, _, _ in cols].index(header) + 1)

def text_block(ws, r, lines, width=100):
    for line in lines:
        if line is None: r += 1; continue
        style, txt = line if isinstance(line, tuple) else ("n", line)
        c = ws.cell(r, 1, txt); c.alignment = WRAP
        c.font = {"t": T, "h": H, "b": B, "m": M}.get(style, N)
        if style == "k": c.font = Font(bold=True, size=13, color=INK); c.fill = KFILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = max(18, 15 * (1 + len(txt) // width))
        r += 1
    return r
def table(ws, r, headers, rows, widths=None, formulas=False):
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h); c.font = B; c.fill = HFILL; c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        r += 1
        for j, v in enumerate(row, 1):
            c = ws.cell(r, j, v); c.border = BOX; c.alignment = WRAP
            if isinstance(v, str) and v.startswith("="): c.font = Font(color=ACC)
    return r + 2

def main(a):
    D = read(os.path.join(a.dir, "COCHRANE_DEMAND_CONSULTANT.csv")); W = read(os.path.join(a.dir, "COCHRANE_WAITLIST_ACTIVITY_CONSULTANT.csv")); E = read(os.path.join(a.dir, "COCHRANE_PLACEMENT_ACTIVITY_CONSULTANT.csv"))
    inc = [p for p in D if p["INCIDENT_DEMAND_SCOPE"] == "1"]; coh = [p for p in inc if p["COHORT"]]; c = Counter(p["COHORT"] for p in coh); dc = Counter(p["D_CLASS"][:2] for p in coh if p["COHORT"] == "D")
    res = c["A"] + c["C"] + c["D"]; n_act = sum(1 for p in D if p["CONSULTANT_ACTIVITY_SCOPE"] == "1")
    wb = Workbook(); wb.remove(wb.active)
    PC, WC, EC = person_cols(), wait_cols(), event_cols()

    # ── START HERE ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("START HERE"); ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH": ws.column_dimensions[col].width = 18
    r = text_block(ws, 1, [
        ("t", "Cochrane Type A/B continuing-care demand, FY2022 to FY2026"),
        ("m", "Five fiscal years, 1 April 2021 to 31 March 2026. Type A = long-term care; Type B = designated supportive living level 4. Cochrane facilities = Bethany Cochrane LTC, Hawthorne SL4, Hawthorne SL4D. Big Hill Lodge is outside this extract. Fiscal years are named by the year they end: FY2023 is April 2022 to March 2023."),
        None, ("h", "Two populations")])
    r = table(ws, r, ["Population", "Definition", "People", "Field in the client list"], [
        ["1. New demand in the five years", "People whose need for a Type A/B bed first arose between April 2021 and March 2026 and who have a link to Cochrane. The cohorts A, B, C and D are defined within this population only.", len(inc), "New demand FY2022-FY2026? = Yes"],
        ["2. Cochrane-related activity in the five years", "People who, at any time in the five years, rated a Cochrane site on the waitlist, were on the waitlist while living in the Town of Cochrane, or were admitted to a Cochrane facility, whenever their need first arose.", n_act, "Cochrane activity FY2022-FY2026? = Yes"],
        ["Client list", "Population 1 and Population 2 together, each person once. Two Yes/No columns record membership.", len(D), ""]])
    r = text_block(ws, r, [("h", "The four cohorts (Population 1)")])
    r = table(ws, r, ["Cohort", "Definition", "People", "Note"], [
        ["A", "Town of Cochrane resident; first placed in a Cochrane facility", c["A"], ""],
        ["B", "Not a Town of Cochrane resident (the Cochrane catchment included); placed in a Cochrane facility", c["B"], "Use of Cochrane beds by people from outside the Town. Not part of resident demand."],
        ["C", "Town of Cochrane resident; first placed outside Cochrane", c["C"], "Based on the first placement; later moves into Cochrane appear in PLACEMENT ACTIVITY."],
        ["D", "Town of Cochrane resident; no Type A/B placement observed by 31 March 2026", c["D"], f"D1 still waiting {dc['D1']}; D2 died before placement {dc['D2']}; D3 left the list {dc['D3']}."]])
    r = text_block(ws, r, [("k", f"Town of Cochrane resident demand = A + C + D = {c['A']} + {c['C']} + {c['D']} = {res}. B is reported separately."), None,
        ("h", "Population 1 beyond the four cohorts"),
        f"{len(inc)-len(coh)} people in Population 1 fall outside A to D: they rated a Cochrane site but live elsewhere, or their residence could not be resolved, or they reached a Cochrane facility only after a first placement elsewhere. They remain in the client list with a descriptive entry in the 'Group (in words)' column.", None,
        ("h", "How the sheets fit together"),
        "HOW PEOPLE ARE COUNTED sets out the rules step by step. HEADLINE holds the findings; each figure is a formula over the data sheets, shown beside the value reproduced by the independent reviewer. CLIENT LIST has one row per person, WAITLIST ACTIVITY one row per waitlist spell, PLACEMENT ACTIVITY one row per Type A/B admission. The anonymous Study ID links the three. DATA DICTIONARY defines every column. QA CHECKS recalculates the reconciliation checks from the data sheets.", None,
        ("h", "Scope notes"),
        "D means that no Type A/B placement was recorded in the Calgary and Edmonton placement systems by 31 March 2026; it is not a measure of unmet need across Alberta.   C reflects the first placement only.   'Rated for a Cochrane site' is what the waitlist recorded; whether it represents a stated preference has not been confirmed with the program.   A person can have several waitlist spells and several admissions; the activity sheets count spells and admissions, the client list counts people.   No health number, date of birth, patient system number, address or postal code appears in this workbook."])

    # ── HOW PEOPLE ARE COUNTED ───────────────────────────────────────────────
    ws = wb.create_sheet("HOW PEOPLE ARE COUNTED"); ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH": ws.column_dimensions[col].width = 18
    r = text_block(ws, 1, [("t", "How every person is counted"), ("m", "Every figure in the workbook follows these steps and nothing else."), None, ("h", "Population 1 — new demand, step by step")])
    r = table(ws, r, ["Step", "Rule", "Plain meaning"], [
        ["1. The demand event", "The earlier of: first approval for a Type A/B bed on the waitlist; first admission to a Type A/B bed.", "The day the need was first recorded."],
        ["2. New demand only", "The demand event falls between 1 April 2021 and 31 March 2026, and the person was not already living in residential continuing care at that moment.", "People whose need began earlier are not new demand; if they had Cochrane activity they are in Population 2."],
        ["3. Where the person lived", "Provincial Registry first: the most recent address in the three fiscal years before the demand event, matched to Statistics Canada geography. If nothing usable there, the Strata address in force on the demand date. Otherwise unresolved.", "A postal code decides. A city name never does. An invalid or dummy postal code never does."],
        ["   Town of Cochrane", "Postal code inside the Town's census boundary.", ""], ["   Cochrane catchment", "Surrounding Rocky View County area served from Cochrane.", "Counts as non-Town for the cohorts."], ["   Not a Cochrane-area resident", "Anywhere else.", ""],
        ["4. What happened", "The FIRST Type A/B admission on or after the demand event, in the Calgary and Edmonton placement systems, by 31 March 2026.", "If none is recorded, no placement was observed."],
        ["5. The cohort", "A: Town resident, placed in Cochrane.  B: not a Town resident, placed in Cochrane.  C: Town resident, placed outside Cochrane.  D: Town resident, no placement observed.", "Resident demand = A + C + D."],
        ["6. The D split", "D1 still on the list at 31 March 2026.  D2 died before any placement.  D3 left the list with no placement observed.", "Three different findings, never one word."]])
    r = text_block(ws, r, [("h", "Population 2 — Cochrane-related activity, step by step")])
    r = table(ws, r, ["Rule", "Plain meaning"], [
        ["Any Type A/B waitlist spell in the five years in which a Cochrane or Hawthorne site was rated", "The person was seeking a Cochrane bed."],
        ["Any Type A/B waitlist spell in the five years while living in the Town of Cochrane", "A Town resident was waiting, whenever their need began."],
        ["Any admission to a Cochrane facility in the five years", "A Cochrane bed was used."],
        ["No test on when the need first arose, and no exclusion for prior residential care", "This is activity, not new demand. Attributes for these people are measured at their first activity in the period."]])
    r = text_block(ws, r, [("h", "Three sheets, three things being counted")])
    r = table(ws, r, ["Sheet", "One row per", "Counting rows gives"], [["CLIENT LIST", "person", "people"], ["WAITLIST ACTIVITY", "waitlist spell (a continuous run of days on the Type A/B list)", "entries to the waitlist; a person can enter more than once"], ["PLACEMENT ACTIVITY", "Type A/B admission", "admissions; a person can have several"]])
    text_block(ws, r, [("h", "Other rules worth knowing"),
        "Age is the completed age on the reference date, never today's age.   Date of birth is the value agreed by at least two of three systems (Strata, Provincial Registry, Connect Care).   Sex is from the Provincial Registry, with Connect Care as a labelled fallback.   Origin is the setting on the person's FIRST waitlist day; if two settings were recorded that day and disagree, the origin is Unknown rather than guessed.   'Rated for a Cochrane site' is what the waitlist recorded, not a confirmed preference.   Days to placement is blank for anyone not placed by 31 March 2026; a wait that has not ended is never shown as a placement time."])

    # ── CLIENT LIST / WAITLIST / PLACEMENT (data first so HEADLINE formulas can reference them) ──
    data_sheet(wb, "CLIENT LIST", D, PC, [16, 12, 12, 34, 8, 44, 30, 13, 34, 10, 12, 13, 13, 12, 10, 9, 6, 16, 26, 26, 34, 24, 26, 12, 12, 34, 10, 30, 13, 10, 32, 11, 12, 12, 14, 16])
    data_sheet(wb, "WAITLIST ACTIVITY", W, WC, [16, 10, 13, 10, 13, 10, 10, 22, 26, 10, 13, 12, 12, 12, 24, 40, 8])
    data_sheet(wb, "PLACEMENT ACTIVITY", E, EC, [16, 13, 10, 36, 9, 10, 22, 26, 14, 14, 10, 24, 22, 40, 8])
    CL = "'CLIENT LIST'!"; WL = "'WAITLIST ACTIVITY'!"; PL = "'PLACEMENT ACTIVITY'!"; PRES = "Person's residence"
    nD, nW, nE = len(D) + 1, len(W) + 1, len(E) + 1
    col = lambda h: L(PC, h); wcol = lambda h: L(WC, h); ecol = lambda h: L(EC, h)
    rng = lambda h: f"{CL}${col(h)}$2:${col(h)}${nD}"; wrng = lambda h: f"{WL}${wcol(h)}$2:${wcol(h)}${nW}"; erng = lambda h: f"{PL}${ecol(h)}$2:${ecol(h)}${nE}"
    cif = lambda h, v: f'COUNTIF({rng(h)},"{v}")'
    def cifs(*pairs): return "COUNTIFS(" + ",".join(f'{rng(h)},"{v}"' for h, v in pairs) + ")"

    # ── HEADLINE ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("HEADLINE", 2); ws.column_dimensions["A"].width = 46
    for col_ in "BCDEFGH": ws.column_dimensions[col_].width = 16
    r = text_block(ws, 1, [("t", "Headline findings"), ("m", "Green cells are live formulas over the data sheets. 'Accepted' is the value the independent reviewer reproduced. The last column must read OK everywhere.")])
    def okcol(rowref): return f'=IF(B{rowref}=C{rowref},"OK","CHECK")'
    ws.cell(r, 1, "Key numbers").font = H; r += 1
    hdr = ["Measure", "Live value", "Accepted", "Status"]; rows = [
        ["Town of Cochrane resident new demand (A + C + D)", f"={cif('Cohort','A')}+{cif('Cohort','C')}+{cif('Cohort','D')}", res],
        ["A — Town residents placed in Cochrane", f"={cif('Cohort','A')}", c["A"]], ["C — Town residents placed outside Cochrane", f"={cif('Cohort','C')}", c["C"]], ["D — Town residents, no placement observed", f"={cif('Cohort','D')}", c["D"]],
        ["B — non-residents placed in Cochrane (report separately)", f"={cif('Cohort','B')}", c["B"]],
        ["D1 still waiting / D2 died / D3 left the list", f'={cif("D outcome","D1*")}&" / "&{cif("D outcome","D2*")}&" / "&{cif("D outcome","D3*")}', f"{dc['D1']} / {dc['D2']} / {dc['D3']}"],
        ["Population 1: new demand in the five years", f"={cif('New demand FY2022-FY2026?','Yes')}", len(inc)], ["Population 2: Cochrane-related activity", f"={cif('Cochrane activity FY2022-FY2026?','Yes')}", n_act],
        ["People in the client list", f"=COUNTA({rng('Study ID')})", len(D)], ["Waitlist spells", f"=COUNTA({wrng('Study ID')})", len(W)], ["Type A/B admissions in the placement sheet", f"=COUNTA({erng('Study ID')})", len(E)],
        ["Admissions to Cochrane facilities", f'=COUNTIF({erng("Cochrane facility?")},"Yes")', sum(1 for e in E if e["PLACEMENT_IN_COCHRANE"] == "1")],
        ["First placements of new-demand people (in the placement sheet)", f'=COUNTIF({erng("First placement of a new-demand person?")},"Yes")', sum(1 for e in E if e["IS_FIRST_PLACEMENT_INCIDENT"] == "1")],
        ["People who rated a Cochrane site at least once", f"={cif('Rated for a Cochrane site?','Yes')}", sum(1 for p in D if p["RATED_FOR_COCHRANE_SITE_FLAG"] == "1")]]
    start = r + 1
    r = table(ws, r, hdr, [[m, f, e, okcol(start + i)] for i, (m, f, e) in enumerate(rows)])
    ws.cell(r, 1, "New demand by fiscal year (Population 1; the year the need first arose)").font = H; r += 1
    start = r + 1; yrows = []
    for i, y in enumerate(FYES):
        rr = start + i; exp = Counter(p["COHORT"] for p in coh if p["DEMAND_FYE"] == str(y))
        yrows.append([f"FY{y}", f"={cifs(('Cohort','A'),('Demand fiscal year',y))}", f"={cifs(('Cohort','C'),('Demand fiscal year',y))}", f"={cifs(('Cohort','D'),('Demand fiscal year',y))}", f"=B{rr}+C{rr}+D{rr}", exp["A"] + exp["C"] + exp["D"], f'=IF(E{rr}=F{rr},"OK","CHECK")', f"={cifs(('Cohort','B'),('Demand fiscal year',y))}"])
    tr = start + len(FYES)
    yrows.append(["Total", f"=SUM(B{start}:B{tr-1})", f"=SUM(C{start}:C{tr-1})", f"=SUM(D{start}:D{tr-1})", f"=SUM(E{start}:E{tr-1})", res, f'=IF(E{tr}=F{tr},"OK","CHECK")', f"=SUM(H{start}:H{tr-1})"])
    r = table(ws, r, ["Fiscal year", "A", "C", "D", "Resident demand A+C+D", "Accepted", "Status", "B (separate)"], yrows)
    ws.cell(r, 1, "Who the people are (Population 2, everyone in the client list)").font = H; r += 1
    who = [["Age under 65", f"={cif('Age group','<65')}"], ["Age 65 to 74", f"={cif('Age group','65-74')}"], ["Age 75 to 84", f"={cif('Age group','75-84')}"], ["Age 85 and over", f"={cif('Age group','85+')}"],
           ["Median age at reference date", f"=MEDIAN({rng('Age at reference date')})"], ["Women", f"={cif('Sex','F')}"], ["Men", f"={cif('Sex','M')}"], ["Sex unknown", f'=COUNTBLANK({rng("Sex")})']]
    who += [[f"Lived: {v}", f'={cif("Residence", v)}'] for v in ["Town of Cochrane", "Cochrane catchment (Rocky View area)", "Not a Cochrane-area resident", "Unresolved"]]
    who += [[f"Entered the pathway from: {v}", f'={cif("Origin setting", v)}'] for v in ["Community", "Acute Care", "Assisted Living or Other Continuing Care", "Lodge", "Out of Province", "Other", "Unknown"]]
    r = table(ws, r, ["Measure", "Live value"], who)
    ws.cell(r, 1, "Waitlist entries by fiscal year (WAITLIST ACTIVITY sheet; spells, not people)").font = H; r += 1
    wy = [[f"FY{y}", f'=COUNTIF({wrng("List entry fiscal year")},{y})', f'=COUNTIFS({wrng("List entry fiscal year")},{y},{wrng("Rated a Cochrane site in this spell?")},"Yes")', f'=COUNTIFS({wrng("List entry fiscal year")},{y},{wrng(PRES)},"Town of Cochrane")', f'=COUNTIFS({wrng("List entry fiscal year")},{y},{wrng("Care type at entry")},"Type A")', f'=COUNTIFS({wrng("List entry fiscal year")},{y},{wrng("Care type at entry")},"Type B")'] for y in FYES]
    r = table(ws, r, ["Fiscal year", "Waitlist spells starting", "Spells rating a Cochrane site", "Spells by Town residents", "Type A", "Type B"], wy)
    ws.cell(r, 1, "Admissions to Cochrane facilities by fiscal year (PLACEMENT ACTIVITY sheet; every admission, whoever the person)").font = H; r += 1
    ey = [[f"FY{y}", f'=COUNTIFS({erng("Placement fiscal year")},{y},{erng("Cochrane facility?")},"Yes")'] + [f'=COUNTIFS({erng("Placement fiscal year")},{y},{erng("Facility")},"{s}")' for s in ("CAL - Bethany Cochrane LTC_", "CAL - Hawthorne SL4_", "CAL - Hawthorne SL4D")]
          + [f'=COUNTIFS({erng("Placement fiscal year")},{y},{erng("Cochrane facility?")},"Yes",{erng(PRES)},"Town of Cochrane")', f'=COUNTIFS({erng("Placement fiscal year")},{y},{erng("Cochrane facility?")},"Yes",{erng("First placement of a new-demand person?")},"Yes")'] for y in FYES]
    r = table(ws, r, ["Fiscal year", "Admissions to Cochrane facilities", "Bethany Cochrane LTC", "Hawthorne SL4", "Hawthorne SL4D", "Of which Town residents", "Of which first placements of new-demand people"], ey)
    ws.cell(r, 1, "Days from demand event to first placement (new-demand people who were placed; median and middle half)").font = H; r += 1
    dtp = rng("Days from demand to first placement"); cohr = rng("Cohort")
    def med(cond, pct=None):
        f = f'{pct or "MEDIAN"}(IF(({cohr}{cond})*({dtp}<>""),{dtp}{"" if pct is None else ",%s" % (0.25 if pct=="P25" else 0.75)}))'
        return f.replace("P25(", "PERCENTILE(").replace("P75(", "PERCENTILE(")
    wt_rows = []
    for lab, cond in (("All placed (A, B, C)", '<>""'), ("A", '="A"'), ("B", '="B"'), ("C", '="C"')):
        wt_rows.append([lab, ArrayFormula(f"B{r+1+len(wt_rows)}", "=" + med(cond)), ArrayFormula(f"C{r+1+len(wt_rows)}", "=" + med(cond, "P25")), ArrayFormula(f"D{r+1+len(wt_rows)}", "=" + med(cond, "P75"))])
    for lab, cond in (("Type A placements", "Type A"), ("Type B placements", "Type B")):
        tr_ = rng("First placement type"); i = len(wt_rows)
        wt_rows.append([lab, ArrayFormula(f"B{r+1+i}", f'=MEDIAN(IF(({tr_}="{cond}")*({dtp}<>""),{dtp}))'), ArrayFormula(f"C{r+1+i}", f'=PERCENTILE(IF(({tr_}="{cond}")*({dtp}<>""),{dtp}),0.25)'), ArrayFormula(f"D{r+1+i}", f'=PERCENTILE(IF(({tr_}="{cond}")*({dtp}<>""),{dtp}),0.75)')])
    r = table(ws, r, ["Group", "Median days", "25th percentile", "75th percentile"], wt_rows)
    text_block(ws, r, [("m", "Waits are measured from the demand event to the first observed placement, placed people only. Later demand years are right-censored at 31 March 2026, so recent medians are biased low. The reviewer independently reproduced the 28-day median under both approval-date rules.")])

    # ── DATA DICTIONARY ──────────────────────────────────────────────────────
    ws = wb.create_sheet("DATA DICTIONARY"); ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 40; ws.column_dimensions["C"].width = 34; ws.column_dimensions["D"].width = 90
    r = text_block(ws, 1, [("t", "Data dictionary"), ("m", "Every column of the three data sheets: the plain header, the technical field it comes from, and what it means.")])
    MEAN = {"Study ID": "Anonymous code for the person; the same code in all three sheets. Cannot be turned back into a health number.",
        "New demand FY2022-FY2026?": "Yes = Population 1: the person's need for a Type A/B bed first arose between April 2021 and March 2026 (technical: INCIDENT_DEMAND_SCOPE). The cohorts A to D exist only for these people.",
        "Cochrane activity FY2022-FY2026?": "Yes = Population 2: the person rated a Cochrane site, was on the waitlist as a Town resident, or was admitted to a Cochrane facility in the five years (technical: CONSULTANT_ACTIVITY_SCOPE).",
        "Why the person is in the list": "New demand; approved just before the window with activity in it (carry-in); already in residential care before April 2021; or earlier/non-new demand with activity in the period.",
        "Cohort": "A, B, C or D as defined on the START HERE sheet; blank for anyone outside the four cohorts.", "Group (in words)": "The cohort in words, or the descriptive group for people outside A to D.",
        "D outcome": "For D only: D1 still waiting at 31 March 2026; D2 died before any placement; D3 left the list with no placement observed.",
        "Reference date": "The date the person's age, residence and origin were measured at: the demand event for Population 1; the first activity in the period for Population-2-only people.",
        "What the reference date is": "Says which of the two the reference date is.", "Demand fiscal year": "Fiscal year of the demand event (FY2023 = April 2022 to March 2023). Population 1 only.",
        "Demand event type": "'approval' if the demand event was a waitlist approval; 'admission' if the person was admitted with no recorded approval.",
        "First waitlist date": "First day on the Type A/B waitlist in the period.", "Last seen on waitlist": "Last day on the waitlist in the period.", "Still on list at 31 Mar 2026?": "Yes = on the waitlist on the last day of the period.",
        "Age at reference date": "Completed years on the reference date (birthday test), never today's age.", "Age group": "<65, 65-74, 75-84, 85+.", "Sex": "F or M.", "Sex source": "Provincial Registry, or Connect Care as a labelled fallback when the Registry has no record.",
        "Residence": "Town of Cochrane; Cochrane catchment; Not a Cochrane-area resident; Unresolved. Decided by postal code (Registry first, then Strata).", "Residence community": "The municipality of the deciding postal code (Statistics Canada census subdivision). For an out-of-province address the Strata city is shown and labelled.",
        "Residence address basis": "Whether the address is a Registry fiscal-year address (and which year) or a Strata address version in force at the reference date. A Registry address is not an exact address on the demand date.",
        "Origin setting": "The setting the person entered the pathway from on their first waitlist day: Acute Care, Community, Assisted Living or Other Continuing Care, Lodge, Out of Province, Other, Unknown.", "Origin detail": "The finer category (own home, acute hospital, sub-acute or transition unit, supportive living, long-term care, ...).",
        "Origin tied on entry day?": "Yes = two different locations were recorded on the first day. If both map to the same setting that setting is used; otherwise Unknown, never a guess.",
        "Rated for a Cochrane site?": "Yes = a Cochrane or Hawthorne site appeared among the sites the person was rated for while on the list. This is what the waitlist records; it has not been confirmed as a preference.", "Cochrane sites rated": "Which Cochrane sites.", "Number of sites rated": "How many different sites were rated.",
        "Most often observed rated site": "The site appearing most often across the person's daily waitlist records. Because records are daily, a long-lasting rating appears more often: an observation, not a preference.",
        "First placement date": "First Type A/B admission on or after the demand event, by 31 March 2026. Population 1 only.", "Placement fiscal year": "Fiscal year of that admission.", "First placement site": "The facility.", "First placement type": "Type A or Type B.",
        "First placement in Cochrane?": "Yes = the first placement was in a Cochrane facility.", "Any Cochrane admission in period?": "Yes = any admission to a Cochrane facility in the five years, first or later.",
        "Days from demand to first placement": "Days from the demand event to the first placement. Blank for anyone not placed by 31 March 2026.", "Days waiting as of 31 Mar 2026 (still waiting only)": "For people still on the list at the end of the period: days waited so far. A censored duration, not a placement time.",
        "Spell number for this person": "1 for the person's first waitlist spell in the period, 2 for the second, and so on.", "List entry date": "First day of the spell.", "List entry fiscal year": "Fiscal year the spell began; count spells by this for entries per year.", "Last seen on list": "Last day of the spell.", "Days observed in spell": "Number of census days in the spell.",
        "Care type at entry": "Type A or Type B at the start of the spell.", "First approval date in spell": "First assessment approval recorded in the spell.", "Rated a Cochrane site in this spell?": "Yes = a Cochrane or Hawthorne site was rated during the spell.", "Already on list at 1 Apr 2021?": "Yes = the spell was already open when the period began.",
        "Person's residence": "Copied from the client list.", "Person's cohort": "Copied from the client list.", "Admission date": "The admission date.", "Facility": "Where the person was admitted.", "Care type": "Type A or Type B.", "Cochrane facility?": "Yes = Bethany Cochrane LTC, Hawthorne SL4 or Hawthorne SL4D.",
        "Admitted from (setting)": "The setting the person was admitted from, from the admission record.", "Admitted from (detail)": "The finer category.", "First placement of a new-demand person?": "Yes = this admission is the first placement of a Population-1 person, the one that decides the cohort. Sums to the number of Population-1 people who were placed.",
        "Person's first admission in the period?": "Yes = the person's first Type A/B admission in the five years, whoever they are.", "Admission number for this person": "1 for the person's first admission in the period, 2 for the second, and so on.", "Person's residence community": "Copied from the client list."}
    rows = []
    for sheet, cols in (("CLIENT LIST", PC), ("WAITLIST ACTIVITY", WC), ("PLACEMENT ACTIVITY", EC)):
        for h, tech, _ in cols: rows.append([sheet, h, tech, MEAN.get(h, "")])
    table(ws, r, ["Sheet", "Column", "Technical field", "Meaning"], rows)

    # ── QA CHECKS ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("QA CHECKS"); ws.column_dimensions["A"].width = 64; ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 14; ws.column_dimensions["D"].width = 10
    r = text_block(ws, 1, [("t", "QA checks"), ("m", "Every 'Actual' cell is a formula over the data sheets. If a sheet is edited, these will say so.")])
    checks = [("Client list rows", f"=COUNTA({rng('Study ID')})", len(D)), ("Unique study IDs in the client list", f"=SUMPRODUCT(1/COUNTIF({rng('Study ID')},{rng('Study ID')}))", len(D)),
              ("Cohort A", f"={cif('Cohort','A')}", c["A"]), ("Cohort B", f"={cif('Cohort','B')}", c["B"]), ("Cohort C", f"={cif('Cohort','C')}", c["C"]), ("Cohort D", f"={cif('Cohort','D')}", c["D"]),
              ("Resident demand A + C + D", f"={cif('Cohort','A')}+{cif('Cohort','C')}+{cif('Cohort','D')}", res), ("D1 + D2 + D3 = D", f'={cif("D outcome","D1*")}+{cif("D outcome","D2*")}+{cif("D outcome","D3*")}', c["D"]),
              ("Every A, C, D is a Town of Cochrane resident (count of exceptions)", f'={cifs(("Cohort","A"),("Residence","<>Town of Cochrane"))}+{cifs(("Cohort","C"),("Residence","<>Town of Cochrane"))}+{cifs(("Cohort","D"),("Residence","<>Town of Cochrane"))}', 0),
              ("Every B is not a Town resident (count of exceptions)", f'={cifs(("Cohort","B"),("Residence","Town of Cochrane"))}', 0),
              ("Every A and B first placed in Cochrane (count of exceptions)", f'={cifs(("Cohort","A"),("First placement in Cochrane?","No"))}+{cifs(("Cohort","B"),("First placement in Cochrane?","No"))}', 0),
              ("Every C first placed outside Cochrane (count of exceptions)", f'={cifs(("Cohort","C"),("First placement in Cochrane?","Yes"))}', 0),
              ("Every D has no first placement date (count of exceptions)", f'=COUNTIFS({rng("Cohort")},"D",{rng("First placement date")},"<>")', 0),
              ("Every cohort member is in Population 1 (count of exceptions)", f'=COUNTIFS({rng("Cohort")},"<>",{rng("New demand FY2022-FY2026?")},"No")', 0),
              ("Waitlist spells", f"=COUNTA({wrng('Study ID')})", len(W)), ("Type A/B admissions", f"=COUNTA({erng('Study ID')})", len(E)), ("Admissions to Cochrane facilities", f'=COUNTIF({erng("Cochrane facility?")},"Yes")', sum(1 for e in E if e["PLACEMENT_IN_COCHRANE"] == "1")),
              ("First placements of new-demand people = Population-1 people with a first placement date", f'=COUNTIF({erng("First placement of a new-demand person?")},"Yes")', sum(1 for p in inc if p["FIRST_PLACEMENT_DT"])),
              ("Placement rows flagged as a new-demand first placement for a person who is not new demand", f'=COUNTIFS({erng("First placement of a new-demand person?")},"Yes",{erng("Why the person is in the list")},"<>New demand in FY2022-FY2026")', 0),
              ("Days to placement present for someone with no placement date (count)", f'=COUNTIFS({rng("Days from demand to first placement")},"<>",{rng("First placement date")},"")', 0)]
    start = r + 1
    r = table(ws, r, ["Check", "Actual", "Expected", "Result"], [[lab, f, e, f'=IF(ROUND(B{start+i},6)=ROUND(C{start+i},6),"PASS","FAIL")'] for i, (lab, f, e) in enumerate(checks)])
    ws.cell(r, 1, "Overall").font = B; ws.cell(r, 2, f'=IF(COUNTIF(D{start}:D{start+len(checks)-1},"FAIL")=0,"ALL PASS","FAIL: "&COUNTIF(D{start}:D{start+len(checks)-1},"FAIL"))').font = Font(bold=True, color=ACC)

    wb.save(a.out); print("wrote", a.out)

if __name__ == "__main__":
    ap = argparse.ArgumentParser(); ap.add_argument("--dir", default="deliverables"); ap.add_argument("--out", default="deliverables/COCHRANE_CLIENT_LIST_EXPLAINED.xlsx"); main(ap.parse_args())
